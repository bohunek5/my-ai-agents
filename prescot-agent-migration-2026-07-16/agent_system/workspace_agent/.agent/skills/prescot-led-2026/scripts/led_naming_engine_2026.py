import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
import re
import json

# DEFINITIVE SOURCE MAPPING - BASED ON ACTUAL FILE ANALYSIS
SOURCE_CORRUPT_MAP = {
    '\u012e': 'ż', # Į -> ż (lowercase requested)
    '\u016f': 'ó', # ů -> ó
    '\u2265': 'ł', # ≥ -> ł
    '\u0145': 'ż', # Ņ -> ż 
    '\u016e': 'ń', # Ů -> ń 
    '\u00fa': 'ś', # ú -> ś
    '\u00cd': 'ę', # Í -> ę
    '\u013b': 'ń', # ļ -> ń
    '\u2021': 'ą', # ‡ -> ą
    '\u00f8': 'ś', # ø -> ś
    '\u00ca': 'ę', # Ê -> ę
    '\u0105': 'ą',
    '≥': 'ł'
}

def fix_polish_chars_v56(text):
    if not isinstance(text, str) or text == 'nan': return text
    
    # 1. Manual string level fixes - FORCE LOWERCASE COLORS
    text = text.replace('Įů≥ta', 'żółta').replace('įů≥ta', 'żółta')
    text = text.replace('RůŅowa', 'różowa').replace('růŅowa', 'różowa')
    text = text.replace('PomaraŮczowa', 'pomarańczowa')
    text = text.replace('Bia≥a', 'biała').replace('ciep≥a', 'ciepła').replace('zimna', 'zimna')
    text = text.replace('Czerwona', 'czerwona').replace('Niebieska', 'niebieska').replace('Zielona', 'zielona')
    text = text.replace('úwiat≥a', 'światła')
    
    # 2. Character by character replacement
    res = []
    for c in text:
        res.append(SOURCE_CORRUPT_MAP.get(c, c))
    text = "".join(res)
    
    # 3. Final sanitization
    text = text.replace('Led', 'LED').replace('led', 'LED')
    text = text.replace('Tasma', 'Taśma').replace('Ta\u015bma', 'Taśma')
    
    return text.lower() if len(text) < 20 and any(c in text for c in ['biała', 'żółta', 'czerwona', 'niebieska', 'zielona']) else text

def final_polish_name_v56(name):
    if not name: return ""
    # Force RGB variants to uppercase
    name = re.sub(r'(?i)rgb\+cct', 'RGB+CCT', name)
    name = re.sub(r'(?i)rgbw', 'RGBW', name)
    name = re.sub(r'(?i)\brgb\b', 'RGB', name)
    
    # Lowercase color names in name
    colors = ['Biała', 'Ciepła', 'Zimna', 'Neutralna', 'Żółta', 'Różowa', 'Czerwona', 'Zielona', 'Niebieska', 'Pomarańczowa']
    for c in colors:
        name = name.replace(c, c.lower())
    
    # Deduplicate RGB variants
    if 'RGB+' in name or 'RGBW' in name:
        name = name.replace(' RGB ', ' ')
    
    parts = name.split()
    seen = set()
    deduped = []
    for p in parts:
        p_low = p.lower()
        if p_low in seen and p_low in ['rgb', 'rgbw', 'rgb+cct', 'led', 'taśma']:
            continue
        deduped.append(p)
        seen.add(p_low)
    
    name = " ".join(deduped)
    name = name.replace('LED/M', 'led/m').replace('W/M', 'W/m').replace('LM/M', 'lm/m')
    return re.sub(r'\s+', ' ', name).strip()

def extract_number(val):
    if pd.isna(val) or str(val).lower() == 'nan' or str(val).strip() == '': return ""
    s = str(val).strip().replace(',', '.')
    match = re.search(r'(\d+\.?\d*)', s)
    if match:
        num = match.group(1).rstrip('.')
        if num.endswith('.0'): num = num[:-2]
        return num
    return ""

def build_symbol_decoder(excel_path):
    mapping = {}
    try:
        df5 = pd.read_excel(excel_path, sheet_name='Arkusz5')
        for _, row in df5.iterrows():
            v, den, wi, pwr, cri, br = [extract_number(row.get(k)) for k in ['napięcie', 'led/m', 'pcb', 'moc', 'CRI', 'lumeny 4K']]
            for col in df5.columns:
                val = str(row.get(col)).strip()
                if '-' in val and len(val) > 5:
                    mapping[val.lower()] = {"V": v, "Density": den, "Width": wi, "Power": pwr, "CRI": cri, "Brightness": br}
        df_up = pd.read_excel(excel_path, sheet_name='Uporządkowane')
        for _, row in df_up.iterrows():
            idx = str(row.get('INDEKS_HANDLOWY')).lower().strip()
            if idx and idx != 'nan':
                mapping[idx] = {
                    "V": extract_number(row.get('Napięcie')),
                    "Density": extract_number(re.search(r'(\d+)led', str(row.get('Typ diody'))).group(1)) if re.search(r'(\d+)led', str(row.get('Typ diody'))) else "",
                    "Power": extract_number(row.get('Moc')), "CRI": extract_number(row.get('CRI')), "Brightness": extract_number(row.get('Jasność')),
                    "Type": str(row.get('Typ diody')).strip(), "Color": str(row.get('Barwa')).strip(), "Length": extract_number(row.get('Rolka'))
                }
    except Exception as e: print(f"Decoder error: {e}")
    return mapping

def transform_logic():
    input_path = os.path.expanduser('~/Downloads/Taśmy nowa nazwa.xlsx')
    output_path = os.path.expanduser('~/Downloads/LED_MIAZGA_V5_6_FINAL_FINAL.xlsx')
    
    symbol_decoder = build_symbol_decoder(input_path)
    try:
        with open('/tmp/super_master_specs.json', 'r') as f: pdf_data = json.load(f)
    except: pdf_data = {}
    
    xl = pd.ExcelFile(input_path)
    df_sys = pd.read_excel(xl, sheet_name='Dane z Systemu')
    df_econ = pd.read_excel(xl, sheet_name='economic')
    df_delux = pd.read_excel(xl, sheet_name='delux')
    df_del_prem = pd.read_excel(xl, sheet_name='delux i premium')
    df_target = pd.read_excel(xl, sheet_name='taśmy system')
    
    def fmt_ean(val):
        s = str(val).strip().split('.')[0]
        return s if s.isdigit() and len(s) >= 10 else ""

    master_map = {fmt_ean(row.get('KOD_KRESKOWY')): row.to_dict() for _, row in df_sys.iterrows() if fmt_ean(row.get('KOD_KRESKOWY'))}
    series_map = {fmt_ean(row.get('EAN')): "Standard" for _, row in df_econ.iterrows()}
    for _, row in df_delux.iterrows(): series_map[fmt_ean(row.get('EAN'))] = "Delux"
    for _, row in df_del_prem.iterrows():
        e = fmt_ean(row.get('EAN'))
        if e: series_map[e] = "Premium" if 'premium' in str(row.get('Nazwa', '')).lower() else "Delux"

    aliases = {'Kod kreskowy': 'EAN', 'Nazwa ca≥a': 'Nazwa', 'NapiÍcie Wejúciowe': 'V', 'Barwa úwiat≥a': 'Barwa', 'Jasnoúś': 'Brightness', 'Pobůr mocy': 'Power', 'Iloúś diod': 'Density', 'Wymiar': 'Width'}
    df_target = df_target.rename(columns=lambda x: aliases.get(x, x))
    
    final_results = []
    stats = {'total': 0, 'green': 0, 'yellow': 0, 'red': 0}
    VALID_LENGTHS = {'1', '5', '10', '25', '40', '50', '100'}

    def smart_find_local(text_blob, pattern_regex):
        if not text_blob: return ""
        match = re.search(pattern_regex, str(text_blob).lower())
        if match: return match.group(1).rstrip('.')
        return ""

    for _, row in df_target.iterrows():
        ean = fmt_ean(row.get('EAN'))
        master = master_map.get(ean, {})
        pdf_item = pdf_data.get(ean, {})
        ci = str(row.get('Indeks handlowy','')).lower().strip()
        ck = fix_polish_chars_v56(str(row.get('Indeks katalogowy',''))).strip()
        decoded = symbol_decoder.get(ci, {})
        
        name_clean = fix_polish_chars_v56(str(row.get('Nazwa', '')))
        search_blob = f"{name_clean.lower()} | {ci}"
        series = series_map.get(ean)
        if not series: series = "Premium" if 'premium' in search_blob else ("Delux" if 'delux' in search_blob else "Standard")

        def get_p(lk, mk, pdf_k, dec_k, rx):
            v = decoded.get(dec_k, "")
            if not v: v = extract_number(row.get(lk))
            if not v: v = extract_number(master.get(mk))
            if not v: v = pdf_item.get(pdf_k, "")
            if not v: v = smart_find_local(search_blob, rx)
            return str(v)

        v = get_p('V', 'Napiecie_wejsciowe', 'V', 'V', r'(\d+)v')
        den = get_p('Density', 'Ilosc_diod', 'Density', 'Density', r'(\d+)led')
        pwr_str = get_p('Power', 'Moc', 'Moc', 'Power', r'(\d+\.?\d*)w')
        
        # Suffix Extraction Logic
        color_override, length_override, width_override = None, None, None
        parts = ci.split('-')
        last_part = parts[-1] if parts else ""
        
        if len(last_part) == 4 and last_part.isdigit():
            w_candidate, l_candidate = last_part[:2], last_part[2:]
            if w_candidate in {'4', '5', '6', '8', '10', '12', '08', '05', '04', '06'}:
                width_override = str(int(w_candidate))
                l_v = str(int(l_candidate))
                if l_v in VALID_LENGTHS: length_override = l_v
        
        if last_part == '2750': color_override, length_override = "2700K", "50"
        elif last_part == '275': color_override, length_override = "2700K", "5"
        elif last_part == '1250': width_override, length_override = "12", "50"
        elif last_part == '0850' or (last_part == '850' and not width_override): width_override, length_override = "8", "50"
        elif last_part == '10': length_override = "10"
        
        if ci.endswith('27k'): color_override = "2700K"
        for l_v in VALID_LENGTHS:
            if ci.endswith(f'ww{l_v}') or ci.endswith(f'nw{l_v}'): length_override = l_v
        
        leng = length_override if length_override else decoded.get('Length', "")
        if not leng and len(parts) > 1 and parts[-1].isdigit() and parts[-1] in VALID_LENGTHS: leng = parts[-1]
        if not leng: leng = extract_number(row.get('Long'))
        if not leng: leng = extract_number(master.get('Dlugosc'))
        if not leng:
             l_par = re.search(r'\((\d+)\)', search_blob)
             if l_par: leng = l_par.group(1)
        if leng and str(leng) not in VALID_LENGTHS:
            found_ls = [l for l in VALID_LENGTHS if str(leng).endswith(l)]
            leng = max(found_ls, key=len) if found_ls else ""

        # POWER NORMALIZATION (V5.6)
        pwr = float(pwr_str) if pwr_str and pwr_str.replace('.', '').isdigit() else 0
        if series == "Standard" and pwr > 28 and leng and str(leng).isdigit():
            pwr = round(pwr / float(leng), 1)
        pwr_final = str(pwr) if pwr > 0 else pwr_str
        if pwr_final.endswith('.0'): pwr_final = pwr_final[:-2]

        wi = width_override if width_override else decoded.get('Width', "")
        if not wi:
            for p in parts:
                if p in {'4', '5', '6', '8', '10', '12'}: 
                    if p == leng and parts.index(p) == len(parts)-1: continue
                    wi = p; break
        if not wi: wi = get_p('Width', 'Wymiar', 'Width', 'Width', r'(\d+)mm')

        c_raw = fix_polish_chars_v56(str(row.get('Barwa', ''))).lower()
        if not c_raw or c_raw == 'nan': c_raw = str(decoded.get('Color', '')).lower()
        
        prefix = ""
        if "rgb+cct" in search_blob: prefix = "RGB+CCT"
        elif "rgbw" in search_blob: prefix = "RGBW"
        elif "rgb" in search_blob: prefix = "RGB"

        rules = {'biała ciepła': '3000K', 'ciepła': '3000K', 'ww': '3000K', 'neutralna': '4000K', 'nw': '4000K', 'zimna': '6000K', 'cw': '6000K', 'w': '6000K'}
        
        if color_override: color_final = color_override.lower()
        else:
            k_found = re.search(r'([1-9]\d{3,4})k', search_blob + " " + c_raw)
            if not k_found:
                k_found = re.search(r'\b(10|15)k\b', search_blob + " " + c_raw)
                color_val = f"{k_found.group(1)}000K" if k_found else next((val for k, val in rules.items() if k in search_blob or k in c_raw), (c_raw.strip().lower() if c_raw != 'nan' else ""))
            else:
                val_k = k_found.group(1)
                color_val = f"{val_k}K" if int(val_k) <= 15000 else next((val for k, val in rules.items() if k in search_blob or k in c_raw), "")
            
            color_val = re.sub(r'(?i)rgb', '', color_val).strip()
            color_final = f"{prefix} {color_val}".strip() if prefix else color_val.strip()

        type_raw = str(decoded.get('Type', '')).strip()
        if not type_raw or 'nan' in type_raw.lower(): type_raw = str(master.get('Typ_diody', '')).strip()
        type_clean = type_raw.upper().replace('LED/M', '').replace('led/m', '').replace('LED/m', '').strip()
        if den and den in type_clean: type_clean = type_clean.replace(den, '').strip()
        if not type_clean or type_clean in ['LED', 'NAN']: type_final = 'COB' if 'cob' in search_blob else ('SMD5050' if '5050' in search_blob else 'SMD2835')
        else: type_final = f"SMD{type_clean}" if type_clean.isdigit() else type_clean

        ip = next((r for r in ['IP65', 'IP67', 'IP68'] if r.lower() in search_blob), "IP20")
        br, cr = get_p('Brightness', 'Jasnosc', 'Lumen', 'Brightness', r'(\d+)lm'), decoded.get('CRI', "")
        if not cr: cr = extract_number(master.get('CRI'))
        if not cr: cr = pdf_item.get('Ra', "")
        if not cr:
            f = re.search(r'(?:ra|cri)\s?(\d+)', search_blob)
            if f: cr = f.group(1)

        series_name = series if series else "Standard"
        warr = "2Y" if series_name == "Standard" else ("PL5Y" if series_name == "Premium" else "PL7Y")
        if "cob" in type_final.lower() or "cob" in search_blob: warr = "3Y"
        
        p_n = ["Taśma LED", series_name, f"{v}V" if v else "", f"{den}led/m" if den else "", type_final, f"{pwr_final}W/m" if pwr_final else "", color_final, ip, f"{br}lm/m" if br else "", f"{wi}mm" if wi else "", f"CRI{cr}" if cr else "", warr, f"{leng}m" if leng else ""]
        final_name = final_polish_name_v56(" ".join([p for p in p_n if p and p.strip()]))
        
        missing = [lb for lb, val in zip(["V", "led/m", "W/m", "Barwa", "lm/m", "Width", "CRI"], [v, den, pwr_final, color_final, br, wi, cr]) if not val]
        braki_str = ", ".join(missing) if missing else "KOMPLET"
        score = sum(1 for m in [v, den, pwr_final, color_final, br, wi, cr] if m) / 7
        
        stats['total'] += 1
        if score >= 1.0: stats['green'] += 1
        elif score >= 0.8: stats['yellow'] += 1
        else: stats['red'] += 1
        
        final_results.append([ck, ci.upper(), str(ean), final_name, braki_str, name_clean, score])

    df_final = pd.DataFrame(final_results, columns=['Indeks Katalogowy', 'Indeks Handlowy', 'EAN', 'Nowa Nazwa', 'BRAKI', 'Stara Nazwa', 'Score'])
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Taśmy - OSTATECZNIK')
        ws = writer.sheets['Taśmy - OSTATECZNIK']
        for i, rowdata in enumerate(final_results, start=2):
            s = rowdata[6]
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            if s >= 1.0: fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            elif s >= 0.8: fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws.cell(row=i, column=4).fill = fill
            
        ws.cell(row=1, column=9).value = "PODSUMOWANIE ILOŚCIOWE"
        ws.cell(row=1, column=9).font = Font(bold=True, size=12)
        ws.cell(row=2, column=9).value = f"RAZEM ARTYKUŁÓW: {stats['total']}"
        ws.cell(row=3, column=9).value = f"100% KOMPLETNE: {stats['green']}"
        ws.cell(row=4, column=9).value = f"OPARTE NA SYMBOLU: {stats['yellow']}"
        ws.cell(row=5, column=9).value = f"BRAKI DANYCH: {stats['red']}"

    print(f"MASTER V5.6 DONE: {output_path}")

if __name__ == "__main__":
    transform_logic()
