#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Prescot LED – Living Database API Server (Port: 11435)
Bridges local preview HTML, local Ollama LLM, WAPRO Excel sheet, and source templates.
"""

import os
import re
import sys
import json
import urllib.request
import urllib.error
import subprocess
from http.server import HTTPServer, BaseHTTPRequestHandler

EXCEL_PATH = "/Users/karolbohdanowicz/Downloads/EksportowaneArtykuly.xlsx"
DOWNLOADS_DIR = "/Users/karolbohdanowicz/Downloads"

REBUILD_SCRIPT = "/Users/karolbohdanowicz/.gemini/antigravity-ide/brain/b8eb29db-e654-4804-822a-d909fdf8954e/scratch/rebuild_with_three_tabs.py"

def query_ollama(model_name, prompt):
    url = "http://localhost:11434/api/chat"
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Jesteś profesjonalnym copywriterem SEO i ekspertem branży oświetleniowej Prescot LED.\n"
                    "Otrzymujesz opis produktu. Twoim zadaniem jest napisać go zupełnie od nowa w języku polskim, "
                    "tak aby zachować wszystkie parametry techniczne (napięcie, moc, barwa, wymiary itp.), "
                    "ale zmienić strukturę zdań i słownictwo w celu uzyskania najwyższej unikatowości SEO "
                    "i uniknięcia duplikatów. Zachowaj dokładnie strukturę tagów HTML (<section>, <h3>, <p>, style, klasy).\n"
                    "ZASADY DLA TAŚM LED:\n"
                    "1. Dla taśm LED o długościach 1m, 3m lub 5m POD ŻADNYM POZOREM NIE UŻYWAJ słów 'rolka', 'szpula', 'rolki' itp. Zamiast tego napisz, że jest to 'wariant cięty z metra' (np. 'wygodny wariant cięty z metra' / 'odcinek cięty z metra').\n"
                    "2. Dla taśm LED o długościach 50m lub 100m (np. modele kończące się na '50' lub '100') podkreślaj, że to 'dłuższy wariant w lepszej cenie' (np. 'szpula zbiorcza w korzystniejszej cenie' / 'dłuższa taśma w lepszej cenie').\n"
                    "Nie zwracaj żadnego wstępu, podsumowania ani znaczników markdown typu ```html. "
                    "Zwróć TYLKO czysty kod HTML z sekcjami."
                )
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        "stream": False
    }
    
    headers = {"Content-Type": "application/json"}
    req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req) as res:
            resp_data = json.loads(res.read().decode("utf-8"))
            return resp_data["message"]["content"].strip()
    except Exception as e:
        print(f"Ollama Error: {e}")
        return None

def update_source_file(brand, model, new_html):
    brand_file_map = {
        "WAPRO": "TOP 21 modeli - opisy KOMPLET ze Scharfer - Prescot LED.html",
        "TIM": "TOP 21 modeli - opisy KOMPLET ze Scharfer - TIM.html",
        "ALLEGRO": "TOP 21 modeli - opisy KOMPLET ze Scharfer - Allegro.html"
    }
    filename = brand_file_map.get(brand.upper())
    if not filename:
        return False
        
    filepath = os.path.join(DOWNLOADS_DIR, filename)
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return False
        
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
        
    pattern = rf'<!--\s*START\s+{model}\s*-->.*?<!--\s*KONIEC\s+{model}\s*-->'
    replacement = f"<!-- START {model} -->\n{new_html}\n<!-- KONIEC {model} -->"
    
    new_content, count = re.subn(pattern, replacement, content, flags=re.DOTALL)
    if count > 0:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"Updated {model} in source file {filename}")
        return True
    else:
        print(f"Could not find block for {model} in {filename}")
        return False

def update_excel_file(model, new_html):
    # Strip blogs for WAPRO Excel description?
    # Usually Excel description matches WAPRO description, but we strip HTML tags or keep them as Excel cells.
    # WAPRO ERP imports raw HTML descriptions, so we write the new HTML description directly.
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found at {EXCEL_PATH}")
        return False
        
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]
        indeks_col = header.index("INDEKS_HANDLOWY") + 1
        opis_col = header.index("OPIS") + 1
        
        updated = False
        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r, column=indeks_col).value
            if val and str(val).strip() == model:
                sheet.cell(row=r, column=opis_col).value = new_html
                updated = True
                break
                
        if updated:
            wb.save(EXCEL_PATH)
            wb.close()
            print(f"Updated Excel row for {model}")
            return True
        else:
            wb.close()
            print(f"Model {model} not found in Excel")
            return False
    except Exception as e:
        print(f"Excel Update Error: {e}")
        return False

def run_rebuild():
    try:
        res = subprocess.run([sys.executable, REBUILD_SCRIPT], capture_output=True, text=True)
        print("Rebuild stdout:", res.stdout)
        return res.returncode == 0
    except Exception as e:
        print(f"Rebuild Subprocess Error: {e}")
        return False

class APIHandler(BaseHTTPRequestHandler):
    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):
        if self.path == '/api/generate':
            self.handle_generate()
        elif self.path == '/api/save':
            self.handle_save()
        else:
            self.send_response(404)
            self.end_headers()

    def handle_generate(self):
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        req = json.loads(post_data.decode('utf-8'))
        
        brand = req.get("brand")
        category = req.get("category")
        model = req.get("model")
        current_html = req.get("current_html", "")
        
        if not brand or not category or not model:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Missing parameters")
            return
            
        print(f"Generation request received for {brand} - {model} ({category})")
        
        # Isolate blog section for WAPRO to satisfy business rules
        blog_section = ""
        content_to_rewrite = current_html
        
        if brand.upper() == "WAPRO":
            sections = re.split(r'(?=<section)', current_html.strip())
            if sections and ("poradnik" in sections[-1].lower() or "poradniki" in sections[-1].lower()):
                blog_section = sections.pop()
                content_to_rewrite = "".join(sections).strip()
                print(f"Isolated blog section for WAPRO model {model} (length: {len(blog_section)} chars)")

        # Pick available Ollama model
        ollama_model = "qwen2.5-coder:7b" # Default
        try:
            with urllib.request.urlopen("http://localhost:11434/api/tags") as res:
                tags = json.loads(res.read().decode('utf-8'))
                avail = [m["name"] for m in tags.get("models", [])]
                if "qwen3.5:latest" in avail:
                    ollama_model = "qwen3.5:latest"
                elif "qwen2.5-coder:7b" in avail:
                    ollama_model = "qwen2.5-coder:7b"
                elif avail:
                    ollama_model = avail[0]
        except Exception:
            pass
            
        # Compile prompt
        prompt = (
            f"Oto opis produktu o indeksie {model} dla kanału {brand} w kategorii {category}:\n"
            f"--------------------------------------------------\n"
            f"{content_to_rewrite}\n"
            f"--------------------------------------------------\n"
            f"Napisz ten opis zupełnie od nowa w języku polskim, tak aby brzmiał świeżo i profesjonalnie, "
            f"ale zawierał identyczne parametry techniczne. Stosuj ten sam układ sekcji i stylów HTML. "
            f"Nie dodawaj żadnych sekcji blogowych na końcu. Zwróć tylko czysty kod HTML bez znaczników markdown."
        )
        
        new_html = query_ollama(ollama_model, prompt)
        if not new_html:
            self.send_response(500)
            self.end_headers()
            self.wfile.write(b"Failed to query local LLM (Ollama)")
            return
            
        # Clean potential markdown codes
        new_html = re.sub(r'```html\s*', '', new_html)
        new_html = re.sub(r'```\s*$', '', new_html)
        new_html = new_html.strip()
        
        # Re-append blog section if isolated
        if blog_section:
            new_html = new_html + "\n\n" + blog_section
            
        # Update files
        update_source_file(brand, model, new_html)
        if brand.upper() == "WAPRO":
            update_excel_file(model, new_html)
            
        # Re-run the compiler script
        rebuild_success = run_rebuild()
        
        response = {
            "success": True,
            "new_html": new_html,
            "rebuild": rebuild_success
        }
        
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(response).encode('utf-8'))

    def handle_save(self):
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        req = json.loads(post_data.decode('utf-8'))
        
        brand = req.get("brand")
        model = req.get("model")
        new_html = req.get("new_html")
        
        if not brand or not model or new_html is None:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Missing parameters")
            return
            
        print(f"Manual save request received for {brand} - {model}")
        
        file_ok = update_source_file(brand, model, new_html)
        excel_ok = True
        if brand.upper() == "WAPRO":
            excel_ok = update_excel_file(model, new_html)
            
        rebuild_ok = run_rebuild()
        
        response = {
            "success": file_ok and excel_ok,
            "rebuild": rebuild_ok
        }
        
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(response).encode('utf-8'))

def serve():
    server_address = ('', 11435)
    httpd = HTTPServer(server_address, APIHandler)
    print("Prescot LED API server running on port 11435...")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    print("Server stopped.")

if __name__ == '__main__':
    serve()
