#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import csv
import ctypes
import hashlib
import json
import re
import smtplib
import ssl
from dataclasses import dataclass
from datetime import date, datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_ID = "1TBkkfo6sHY-LkFYZTJPL2w0nxwfSxwfCQLqqhmQAtmo"
GID = "0"
PUBLIC_CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}"

DOWNLOADS = Path.home() / "Downloads"
STATE_PATH = Path.home() / "Library/Application Support/PrescotLeadMailer/google_sheet_state.json"
LOG_DIR = Path.home() / "Library/Logs/PrescotLeadMailer"
THUNDERBIRD_PROFILE = Path.home() / "Library/Thunderbird/Profiles/1onk9yis.default-release"
THUNDERBIRD_LIBDIR = Path("/Applications/Thunderbird.app/Contents/MacOS")

SMTP_HOST = "prescot.home.pl"
SMTP_PORT = 465
SMTP_USERNAME = "karol.bohdanowicz@prescot.pl"
FROM_NAME = "Karol Bohdanowicz"
TO = ["anna.galor@prescot.pl"]
CC: list[str] = []
TEST_CC = ["karol.bohdanowicz@prescot.pl"]

SELECTED_COLUMNS = [
    (1, "Data dodania"),
    (11, "Platforma"),
    (12, "Imie"),
    (13, "Telefon"),
    (14, "Email"),
    (15, "Firma"),
    (16, "Profil dzialalnosci"),
    (17, "Problem / potrzeba"),
]
OUTPUT_HEADERS = [label for _, label in SELECTED_COLUMNS] + ["Uwagi"]


@dataclass(frozen=True)
class LeadRow:
    values: list[str]
    source_row: int

    @property
    def fingerprint(self) -> str:
        payload = "|".join(normalize_key(value) for value in self.values)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()


class SECItem(ctypes.Structure):
    _fields_ = [("type", ctypes.c_uint), ("data", ctypes.c_void_p), ("len", ctypes.c_uint)]


def log(message: str) -> None:
    print(f"[{datetime.now().isoformat(timespec='seconds')}] {message}", flush=True)


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def clean_phone(value: str) -> str:
    value = str(value or "").strip()
    return value[2:] if value.startswith("p:") else value


def parse_sheet_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(str(value or "").strip())
    except ValueError:
        return datetime.min


def read_state() -> dict:
    if not STATE_PATH.exists():
        return {"seen_hashes": [], "runs": []}
    return json.loads(STATE_PATH.read_text(encoding="utf-8"))


def write_state(state: dict) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def public_csv_bytes() -> bytes:
    request = Request(PUBLIC_CSV_URL, headers={"User-Agent": "PrescotLeadMailer/1.0"})
    with urlopen(request, timeout=35) as response:
        data = response.read()
    sample = data[:512].lower()
    if sample.startswith(b"<!doctype html") or b"document-root" in sample or b"accounts.google.com" in sample:
        raise RuntimeError("Google Sheet zwrocil strone logowania zamiast CSV.")
    return data


def latest_downloaded_csv(allow_stale: bool = False) -> Path | None:
    files = sorted(DOWNLOADS.glob("Prescot leady - Arkusz1*.csv"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not files:
        return None
    latest = files[0]
    if allow_stale:
        return latest
    if datetime.fromtimestamp(latest.stat().st_mtime).date() == date.today():
        return latest
    return None


def load_csv_bytes(input_csv: Path | None, allow_stale_csv: bool) -> bytes:
    if input_csv:
        log(f"Reading local CSV: {input_csv}")
        return input_csv.read_bytes()
    try:
        log("Downloading Google Sheet public CSV export")
        return public_csv_bytes()
    except (HTTPError, URLError, RuntimeError) as error:
        log(f"Public CSV endpoint unavailable: {error}")
    local_csv = latest_downloaded_csv(allow_stale=allow_stale_csv)
    if local_csv:
        log(f"Using downloaded CSV fallback: {local_csv}")
        return local_csv.read_bytes()
    raise RuntimeError(
        "Nie udalo sie pobrac arkusza. Udostepnij Google Sheet do odczytu linkiem albo zapisz swiezy CSV w Pobranych."
    )


def decode_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig")
    return list(csv.reader(text.splitlines()))


def selected_value(row: list[str], index: int) -> str:
    return row[index].strip() if index < len(row) else ""


def extract_leads(rows: list[list[str]]) -> list[LeadRow]:
    leads: list[LeadRow] = []
    for source_row, row in enumerate(rows[1:], start=2):
        if not any(str(cell).strip() for cell in row):
            continue
        values = [selected_value(row, index) for index, _ in SELECTED_COLUMNS]
        values[3] = clean_phone(values[3])
        if not values[0] or not any(values[1:]):
            continue
        leads.append(LeadRow(values=values, source_row=source_row))
    return leads


def sort_leads(leads: list[LeadRow]) -> list[LeadRow]:
    return sorted(leads, key=lambda lead: parse_sheet_datetime(lead.values[0]), reverse=True)


def filter_new_leads(leads: list[LeadRow], state: dict, since: datetime | None = None) -> list[LeadRow]:
    seen = set(state.get("seen_hashes", []))
    out = []
    for lead in leads:
        if lead.fingerprint in seen:
            continue
        if since and parse_sheet_datetime(lead.values[0]) < since:
            continue
        out.append(lead)
    return out


def create_xlsx(leads: list[LeadRow], output_path: Path, title: str, note: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nowe leady"
    ws.append([title] + [""] * (len(OUTPUT_HEADERS) - 1))
    ws.append([note] + [""] * (len(OUTPUT_HEADERS) - 1))
    ws.append(OUTPUT_HEADERS)
    for lead in leads:
        ws.append(lead.values + [""])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_HEADERS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(OUTPUT_HEADERS))
    orange = "DE4B26"
    deep = "1A222C"
    light = "F4F6F8"
    muted = "637083"
    thin = Side(style="thin", color="D7DDE4")

    for cell in ws[1]:
        cell.font = Font(bold=True, size=16, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=deep)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[2]:
        cell.font = Font(size=10, color=muted)
        cell.fill = PatternFill("solid", fgColor=light)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(OUTPUT_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for index, width in enumerate([24, 12, 18, 18, 30, 34, 38, 48, 28], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "A4"

    ref = f"A3:{get_column_letter(len(OUTPUT_HEADERS))}{ws.max_row}"
    table = Table(displayName="NoweLeady", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path, read_only=True).close()


def thunderbird_smtp_password() -> str:
    ctypes.CDLL(str(THUNDERBIRD_LIBDIR / "libmozglue.dylib"), mode=ctypes.RTLD_GLOBAL)
    nss = ctypes.CDLL(str(THUNDERBIRD_LIBDIR / "libnss3.dylib"), mode=ctypes.RTLD_GLOBAL)
    nss.NSS_Init.argtypes = [ctypes.c_char_p]
    nss.NSS_Init.restype = ctypes.c_int
    nss.PK11_GetInternalKeySlot.restype = ctypes.c_void_p
    nss.PK11_FreeSlot.argtypes = [ctypes.c_void_p]
    nss.PK11_CheckUserPassword.argtypes = [ctypes.c_void_p, ctypes.c_char_p]
    nss.PK11_CheckUserPassword.restype = ctypes.c_int
    nss.PK11SDR_Decrypt.argtypes = [ctypes.POINTER(SECItem), ctypes.POINTER(SECItem), ctypes.c_void_p]
    nss.PK11SDR_Decrypt.restype = ctypes.c_int
    nss.SECITEM_ZfreeItem.argtypes = [ctypes.POINTER(SECItem), ctypes.c_int]
    nss.NSS_Shutdown.restype = ctypes.c_int

    if nss.NSS_Init(f"sql:{THUNDERBIRD_PROFILE}".encode()) != 0:
        raise RuntimeError("Nie udalo sie zainicjowac NSS dla profilu Thunderbirda.")
    slot = nss.PK11_GetInternalKeySlot()
    if not slot:
        nss.NSS_Shutdown()
        raise RuntimeError("Nie udalo sie pobrac slotu NSS Thunderbirda.")
    try:
        if nss.PK11_CheckUserPassword(slot, b"") != 0:
            raise RuntimeError("Profil Thunderbirda wymaga hasla glownego.")
        logins = json.loads((THUNDERBIRD_PROFILE / "logins.json").read_text(encoding="utf-8")).get("logins", [])
        smtp_login = next((item for item in logins if item.get("hostname") == f"smtp://{SMTP_HOST}"), None)
        if not smtp_login:
            raise RuntimeError(f"Nie znaleziono hasla SMTP w Thunderbirdzie dla {SMTP_HOST}.")

        def decrypt(value: str) -> str:
            raw = base64.b64decode(value)
            buffer = ctypes.create_string_buffer(raw)
            source = SECItem(0, ctypes.cast(buffer, ctypes.c_void_p), len(raw))
            target = SECItem()
            if nss.PK11SDR_Decrypt(ctypes.byref(source), ctypes.byref(target), None) != 0:
                raise RuntimeError("Nie udalo sie odszyfrowac hasla SMTP.")
            try:
                return ctypes.string_at(target.data, target.len).decode("utf-8")
            finally:
                nss.SECITEM_ZfreeItem(ctypes.byref(target), 0)

        username = decrypt(smtp_login["encryptedUsername"])
        if username.lower() != SMTP_USERNAME.lower():
            raise RuntimeError(f"Login SMTP z Thunderbirda to {username}, oczekiwano {SMTP_USERNAME}.")
        return decrypt(smtp_login["encryptedPassword"])
    finally:
        nss.PK11_FreeSlot(slot)
        nss.NSS_Shutdown()


def send_email(subject: str, body: str, attachment_path: Path, to: list[str], cc: list[str]) -> None:
    message = EmailMessage()
    message["From"] = formataddr((FROM_NAME, SMTP_USERNAME))
    message["To"] = ", ".join(to)
    if cc:
        message["Cc"] = ", ".join(cc)
    message["Subject"] = subject
    message.set_content(body)
    message.add_attachment(
        attachment_path.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment_path.name,
    )
    password = thunderbird_smtp_password()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=45, context=ssl.create_default_context()) as smtp:
        smtp.login(SMTP_USERNAME, password)
        smtp.send_message(message, from_addr=SMTP_USERNAME, to_addrs=to + cc)


def mark_seen(state: dict, all_leads: list[LeadRow], sent_count: int, mode: str, xlsx_path: Path | None) -> None:
    seen = set(state.get("seen_hashes", []))
    seen.update(lead.fingerprint for lead in all_leads)
    state["seen_hashes"] = sorted(seen)
    state.setdefault("runs", []).append(
        {
            "time": datetime.now().isoformat(timespec="seconds"),
            "mode": mode,
            "sheet_rows_seen": len(all_leads),
            "sent_count": sent_count,
            "xlsx": str(xlsx_path) if xlsx_path else "",
        }
    )
    state["runs"] = state["runs"][-80:]
    write_state(state)


def build_output_name(now: datetime, count: int, test: bool) -> str:
    suffix = " TEST" if test else ""
    return f"Anna - {now.strftime('%d.%m')} - nowe leady Google Sheet{suffix} ({count}).xlsx"


def run(args: argparse.Namespace) -> int:
    now = datetime.now()
    csv_data = load_csv_bytes(Path(args.input_csv) if args.input_csv else None, args.allow_stale_csv)
    leads = sort_leads(extract_leads(decode_csv(csv_data)))
    if not leads:
        raise RuntimeError("Arkusz nie zawiera leadow w kolumnach B oraz L-R.")

    state = read_state()
    since = datetime.fromisoformat(args.since) if args.since else None
    comparison_state = {"seen_hashes": []} if args.ignore_state else state
    new_leads = sort_leads(filter_new_leads(leads, comparison_state, since))
    log(f"Sheet leads: {len(leads)}; new leads: {len(new_leads)}")

    if not new_leads:
        if args.mark_seen_when_empty:
            mark_seen(state, leads, 0, args.mode, None)
            log("No new leads; state refreshed.")
        else:
            log("No new leads; no email sent.")
        return 0

    output_path = DOWNLOADS / build_output_name(now, len(new_leads), args.test)
    title = f"Nowe leady Prescot - {now.strftime('%d.%m.%Y')}"
    if args.test:
        title = "[TEST] " + title
    note = (
        f"Zrodlo: Google Sheet {SHEET_ID}, kolumny B oraz L-R. "
        f"Wysylka zawiera tylko nowe rekordy: {len(new_leads)}. "
        "Ostatnia kolumna Uwagi jest pusta do uzupelniania."
    )
    create_xlsx(new_leads, output_path, title, note)
    log(f"XLSX created: {output_path}")

    if args.dry_run:
        log("Dry run: email not sent and state not changed.")
        return 0

    subject = f"Nowe leady Prescot - {now.strftime('%d.%m.%Y')} ({len(new_leads)})"
    if args.test:
        subject = "[TEST] " + subject
    body = (
        "Czesc,\n\n"
        f"w zalaczniku nowe unikatowe leady z arkusza Google Sheet - {len(new_leads)} kontaktow.\n"
        "Zakres: kolumna B oraz L-R. Na koncu dodana jest pusta kolumna Uwagi.\n\n"
        "Pozdrawiam,\n"
        "Karol\n"
    )
    cc = list(CC)
    if args.test:
        cc.extend(address for address in TEST_CC if address not in cc)
    send_email(subject, body, output_path, TO, cc)
    log(f"Email sent to {', '.join(TO)}; cc {', '.join(cc) if cc else '-'}")
    mark_seen(state, leads, len(new_leads), args.mode, output_path)
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Send new Google Sheet leads to Anna.")
    parser.add_argument("--input-csv", help="Use a local CSV instead of downloading Google Sheet.")
    parser.add_argument("--allow-stale-csv", action="store_true", help="Allow fallback to an older downloaded CSV.")
    parser.add_argument("--since", help="Only consider leads at or after this ISO datetime.")
    parser.add_argument("--test", action="store_true", help="Add TEST to subject and CC Karol.")
    parser.add_argument("--dry-run", action="store_true", help="Create XLSX only, do not send or mark state.")
    parser.add_argument("--ignore-state", action="store_true", help="Ignore local sent state, useful for one-off resend.")
    parser.add_argument("--mark-seen-when-empty", action="store_true", help="Refresh state when there are no new leads.")
    parser.add_argument("--mode", default="manual", choices=["manual", "scheduled", "test"])
    return parser.parse_args()


if __name__ == "__main__":
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    raise SystemExit(run(parse_args()))
